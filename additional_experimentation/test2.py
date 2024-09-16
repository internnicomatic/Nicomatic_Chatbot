import os
import time
import openpyxl
from langchain_community.llms import LlamaCpp
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from langchain.prompts import PromptTemplate
from sentence_transformers import SentenceTransformer, util
from langchain.callbacks.manager import CallbackManager
from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler
from langchain.docstore.document import Document
from langchain.text_splitter import RecursiveCharacterTextSplitter
import pandas as pd
import os
from llama_index.core import SimpleDirectoryReader
from transformers import AutoTokenizer
from tqdm import tqdm
from langchain.retrievers import ContextualCompressionRetriever
from langchain.retrievers.document_compressors import FlashrankRerank


directory_path = "C:/Users/s.vijaykumar/Desktop/skanda/bot/dataset/extracted_best"
markdown_data = SimpleDirectoryReader(
    directory_path,
    ).load_data()
## Restrucutre the extracted data and metadata
source_docs = [
    Document(
        page_content=doc.text,
        metadata={
            "source": (
                doc.metadata["file_path"].split("/")[0]
                if len(doc.metadata["file_path"].split("/")) > 1
                else doc.metadata["file_path"]
            )
        },
    )
    for doc in markdown_data
]
# xlsx file as input helpful for finding connector family
def read_multiple_xlsx(folder_path):
    Excel_file = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(folder_path, filename)
            data = pd.read_excel(filepath)
            split_docs = [
                Document(page_content=str(row.to_dict()))
                for index, row in data.iterrows()
            ]
            Excel_file.extend(split_docs)
    return Excel_file
## calling the document loading function
excel_data= read_multiple_xlsx(directory_path)

excel_markdown_data=source_docs+excel_data
## Preprocess the extracted data
MARKDOWN_SEPARATORS = [
    "\n#{1,6} ",
    "```\n",
    "\n\\*\\*\\*+\n",
    "\n---+\n",
    "\n___+\n",
    "\n\n",
    "\n",
    " ",
    "",
    "[",
    "]",
]
## Initialize the tokenizer for chunking of data
tokenizer = AutoTokenizer.from_pretrained("thenlper/gte-small")
## Preporcess and actually chunk the data
text_splitter = RecursiveCharacterTextSplitter.from_huggingface_tokenizer(
    tokenizer,
    chunk_size=200,
    chunk_overlap=20,
    add_start_index=True,
    strip_whitespace=True,
    separators=MARKDOWN_SEPARATORS,
)

## To store the chunked data
docs_processed1 = []
## Variable to store unique texts 
unique_texts = {}
## Splits the documents and removes all the duplicated texts
for doc in tqdm(excel_markdown_data):
    new_docs = text_splitter.split_documents([doc])
    for new_doc in new_docs:
        if new_doc.page_content not in unique_texts:
            unique_texts[new_doc.page_content] = True
            docs_processed1.append(new_doc)
### structure connector info from the xlsx into a dictionary
## This is necessary for later hardcoded portion as kb
connector_families = {}
for document in excel_data:
    family_data = eval(document.page_content)
    family = family_data["Family of connectors"]
    characteristics = {}
    for key, value in family_data.items():
        if key != "Family of connectors":
            key_lower = key.lower().replace(" ", "_")
            if isinstance(value, str) and "," in value:
                characteristics[key_lower] = [x.strip() for x in value.split(",")]
            else:
                characteristics[key_lower] = value
    connector_families[family] = characteristics
## Embeddings the cleaned data and store in vector database
def vectordb_in(texts):
    embeddings = HuggingFaceEmbeddings(
        model_name="sentence-transformers/all-MiniLM-L6-v2",
        model_kwargs={"device": "cpu"},
    )
    db = FAISS.from_documents(texts, embeddings)
    DB_FAISS_PATH = "vectorstore1/db_faiss"
    db.save_local(DB_FAISS_PATH)
    return db
## Prompt template
template = """<s>[INST] <<SYS>>
context: {docs_processed1}
conversation history:
{chat_history} 

You are an advanced conversational bot with the capability to engage users in interactive dialogue and ask a series of insightful questions based on a specified feature list. 
Your primary goal is to identify the most suitable family of connector from a selection of four connector families out of AMM, EMM, CMM, and DMM. 
Each family of connectors shares some common features, you are to ask questions to the user based on their response to get essential information from users. 
Keep asking relevant questionto make an educated guess on the right family of connectors."

quesiton: {question}
<</SYS>>
User: I need a connector which is compatible with AWG28 cable. [/INST] Bot:  Absolutely! There are a total of 4 connectors which is inline with your requirement. Do you need EMI shielding with this connector? 
[INST] User: No, I dont need EMI shielding[/INST] Bot: Okay that narrows it down to 3 connector. What is the pitch size required for your use case?
[INST] User: I need connector which has pitch size of 2mm. [/INST] Bot:Great! The most similar connector which meets all your requirement is out CMM line of connectors
</s>
"""
###handle streaming inputs from the LLM
callback_manager = CallbackManager([StreamingStdOutCallbackHandler()])
###conversational chain
def get_conversation_chain(vectordb):
    mistral_llm = LlamaCpp(
        model_path="C:/Users/s.vijaykumar/Desktop/skanda/bot/models/llama-2-7b-chat.Q8_0.gguf",
        temperature=0.2,
        max_tokens=500,
        top_p=1,
        callback_manager=callback_manager,
        n_ctx=1000,
        f16_kv=True,
        verbose=False,
    )
    # from langchain_ollama import ChatOllama
    # mistral_llm=ChatOllama(model='llama3.1',temperature=0)

    ## Creating a retirever to access vector database and retreive the right nodes
    retriever = vectordb.as_retriever(search_type="similarity", search_kwargs={"k": 4})
    ## Reranker for to rank the retireved documents  
    compressor = FlashrankRerank()
    compression_retriever = ContextualCompressionRetriever(
        base_compressor=compressor, base_retriever=retriever
    )
    ## prompt template 
    CONDENSE_QUESTION_PROMPT = PromptTemplate.from_template(template)
    ## Buffer memory to store the entire conversation history
    memory = ConversationBufferMemory(
        memory_key="chat_history",
        input_key="question",
        return_messages=True,
        output_key="answer",
        document_key="docs_processed1",
        max_length=500,
    )
    ## Conversation chain where all the different elemets like memory prompt and LLM are connected. new code uses LCEL.
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=mistral_llm,
        retriever=compression_retriever,
        condense_question_prompt=CONDENSE_QUESTION_PROMPT,
        memory=memory,
        verbose=False,
        return_source_documents=True,
    )  
    print("conversational chain has been created")
    return conversation_chain

### validation using basic fuzzy logic (cosine similarity)
def output_val(response_answer, source_documents, user_question):
    model = SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2")
    similarity_threshold = 0.5
    src_txt = [doc.page_content for doc in source_documents]
    ## Tensors to spread the embeddings across 3D space 
    ans = model.encode(response_answer, convert_to_tensor=True)
    src = model.encode(src_txt, convert_to_tensor=True)
    cosine_scores = util.pytorch_cos_sim(ans, src)
    wb = openpyxl.load_workbook("history.xlsx")
    sheet = wb.active
    rate = input("\n\nwas this response good? yes or no?\n")
    if rate.lower() == "yes":
        suggest = ""
    else:
        suggest = input("what should have been the response?\n")
    sheet.append([user_question, response_answer, rate, suggest])
    wb.save("history.xlsx")
    if any(score.item() > similarity_threshold for score in cosine_scores[0]):
        return True
    return False


# logic for shortlisting the right family of connector:
def get_connector_family():
    print(
        "\n\n Let's find the best connector for your needs! I need more information to suggest the right connector for you. Help me out with these questions:\n"
    )
    answers = {}
    questions = [
        (
            "What is the user case for which you need this connector for? (ex: Harsh industrial, Military Ground applications, Avionics, Medical, Laboratory equipment or UAV)",
            None,
        ),
        (
            "Do you have a general idea of what you type of connector you might need? say a nanod or microd?",
            None,
        ),
        (
            "What type of contact do you need required? (signal, data, low amperage,low frequency, high frequency)",
            None,
        ),
        (
            "What should be the weight of your connector? Are you looking for something which weighs more or less? ",
            None,
        ),
        ("What is the desired pitch size of the connector? (mm)", None),
        ("Do you require EMI shielding for this connector or no?", None),
    ]
    for question, valid_answers in questions:
        ## Handle I don't know answers. need to implement better logic
        while True:
            answer = input(question + "\n")
            if answer.lower() == "i don't know":
                answers[question] = None
                break
            if valid_answers is None or answer.lower() in [
                a.lower() for a in valid_answers
            ]:
                answers[question] = answer
                break
            else:
                print(
                    "Invalid input. Please choose one of the following:", valid_answers
                )
        ## Add interaction to history
        wb = openpyxl.load_workbook("history.xlsx")
        sheet = wb.active
        sheet.append([question, answer])
        wb.save("history.xlsx")
    matching_families = list(connector_families.keys())
    ## Logic to match families from the kb(excel data) 
    for question, answer in answers.items():
        if answer is None:
            continue
        matching_families = [
            family
            for family in matching_families
            if any(
                (
                    any(answer.lower() in str(c).lower() for c in characteristic)
                    if isinstance(characteristic, list)
                    else answer.lower() in str(characteristic).lower()
                )
                for characteristic in connector_families[family].values()
            )
        ]
    if len(matching_families) == 1:
        return matching_families[0]
    else:
        connector_family = None
    return connector_family


### logic for either to go with LLM for response generation or with LLM or shortlisting of connector using fuzzy logic
def analyze_intent_llm(possible, user_question):
    intent_model = SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2")
    user_embedding = intent_model.encode(user_question, convert_to_tensor=True)
    response_embedding = intent_model.encode(possible, convert_to_tensor=True)
    cosine_score = util.pytorch_cos_sim(user_embedding, response_embedding)
    if cosine_score.item() > 0.6:
        return "find_connector_family"
    else:
        return "general_info"


##main function where input is handled and all the functions are called
vectordb = vectordb_in(docs_processed1)
conversation_chain = get_conversation_chain(vectordb)
connector_family_search = False

## Main loop
def main(user_question):
    print("\n\n\n")
    ## split to check if the user needs a connector or just wants general informaiton
    possible = [
        "I am looking for connector"
        or "I need a connector"
        or "I am searching for the best connector for my use case"
        or "find me a good connector which fits my needs"
        or "recommend me a good connector"
    ]
    ## calling the intent analyer to check which path to chose for the interaction
    intent = analyze_intent_llm(possible, user_question)
    start_time = time.time()
    if intent == "find_connector_family":
        family = get_connector_family() ## call the connector family part (Doesnt depend on LLM to shortlist the right product) 
        if family:
            p = f"Based on your answers, the best connector family for you is {family}."
            print(p)
            t = f"\n Here is more information about the {family} connector.\n"
            print(t)
            q = f"tell me more about nicomatic's {family} range of connector." ## For additional info, This but uses the LLM
            response = conversation_chain.invoke({"question": q})
            source_docs = response['source_documents']
            for i, doc in enumerate(source_docs): ## Find the source documents
                print(f'\nSource Document {i+1}\n')
                print(f'Source Text: {doc.page_content}')
                print(f'Document Name: {doc.metadata["source"]}')
            wb = openpyxl.load_workbook("history.xlsx") ## User experience and feedback
            sheet = wb.active
            rate = input("\n\nwas this response good? yes or no?\n")
            if rate.lower() == "yes":
                suggest = ""
            else:
                suggest = input("what should have been the response?\n")
            sheet.append(["final verdict", p, rate, suggest])
            wb.save("history.xlsx")

        else:
            print(
                "Your requirement is not a standard one, But I assure you we offer extensive customizablity using which we can find the right connector for your use case. Here is alink to connect to your sales representative:"
            )
    elif intent == "general_info":
        print("\n\nbot's response: ")
        response = conversation_chain.invoke({"question": user_question})
        print("\n")
        source_docs = response['source_documents']
        for i, doc in enumerate(source_docs):
            print(f'\nSource Document {i+1}\n')
            print(f'Source Text: {doc.page_content}')
            print(f'Document Name: {doc.metadata["source"]}')
    
        if output_val(response['answer'], response['source_documents'], user_question):
            print("\n\n\nvalidated Question: ", user_question)
            print("validated Answer: ", response['answer'])
        else:
            print("The answer given does not fuzzy match with any documents which were fed to me, but was cooked up with the knowldge I have of them")

    else:
        print(
            "I'm not sure what you're looking for. Can you please rephrase your question?"
        )
    end_time = time.time()
    response_time = end_time - start_time
    print(f"\n\n\nresponse time: {response_time:.2f} seconds\n\n\n")

## Main loop
if __name__ == "__main__":
    while True:
        user_question = input("Enter your question: ")
        if user_question.lower() == "done":
            print("Thanks for the interaction!")
            break
        else:
            main(user_question)
